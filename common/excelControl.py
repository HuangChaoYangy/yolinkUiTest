# -*- coding: utf-8 -*-
# @Time    : 2023/11/10 13:41
# @Author  : 黄朝阳
# @FileName: excelControl
# @Software: PyCharm


import openpyxl, sys, os

sys.path.append(os.getcwd())
from base_dir import owner_backer_path


class DoExcel(object):

    def __init__(self, file_name, sheet_name=None):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_sheet(self):
        self.workbook = openpyxl.load_workbook(self.file_name)
        if self.sheet_name:
            sheet = self.workbook[self.sheet_name]
        else:
            sheet = []
            if len(self.workbook.sheetnames) > 1:
                for n in range(len(self.workbook.sheetnames) - 1):
                    sheet_name = (self.workbook.sheetnames[n + 1])
                    sheet.append(self.workbook[sheet_name])
        return sheet

    def get_case(self, sheet):
        # 实例化文件操作
        if type(sheet) != list:
            list_cases = []  # 测试数据
            max_row = sheet.max_row  # 获取最大行
            max_column = sheet.max_column  # 获取最大列
            # 获取 表格数据
            for r in range(2, max_row + 1):  # 遍历行
                # dict_case = {}
                list_case = []
                for j in range(1, max_column + 1):
                    # key = sheet.cell(row=1, column=j).value  # 遍历列
                    # dict_case[key] = sheet.cell(row=r, column=j).value
                    list_case.append(sheet.cell(row=r, column=j).value)
                list_cases.append(list_case)  # 将一行数据放到列表
            # self.save_close_file()
            return list_cases

    def write_result(self, row, actual_result, expect_result, is_pass):
        # 实例化文件操作
        self.workbook = openpyxl.load_workbook(self.file_name)
        # 打开sheet表单
        sheet = self.workbook[self.sheet_name]
        sheet.cell(row=row, column=10).value = actual_result
        sheet.cell(row=row, column=11).value = expect_result
        sheet.cell(row=row, column=12).value = is_pass
        self.save_close_file()

    # 关闭和保存文件
    def save_close_file(self):
        self.workbook.save(self.file_name)
        # self.workbook.close()


if __name__ == "__main__":

    de = DoExcel(file_name=owner_backer_path, sheet_name='sport_params')
    print(de.get_case(de.get_sheet()))
